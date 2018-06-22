import collections
import pandas as pd
import csv
import numpy as np
from openpyxl import load_workbook, Workbook

stop_words = ["and","an","family","restaurant","the","cafe","caffe","pub","bar"]
def match_sentences(sen1,sen2):
    a=0
    percentage_list2 = 0
    percentage_list1 = 0
    str1 = str(sen1)
    str2 = str(sen2)
    str1 = str1.lower().strip()
    str2 = str2.lower().strip()
    list1 = list(set(str1.split(" ")))
    list2 = list(set(str2.split(" ")))
    combined = list1 + list2
    counter = collections.Counter(combined)
    for i in stop_words:
        if(i in counter):
            counter.pop(i)
    for i in counter.values():
        if i>1:
            a =a + 1
    percentage_list1=(a/len(list1))*100
    percentage_list2=(a/len(list2))*100
    return percentage_list1,percentage_list2
def clean_numbers(number):
    number = number.strip(" ")
    number = number.replace("-", "")
    number = number.replace("(", "")
    number = number.replace(")", "")
    number = number.replace("+", "")
    number = number.replace(" ", "")
    return number

def match_numbers(num1,num2):
    percentage = 0
    a = 0
    num1 = str(num1)
    num2 = str(num2)
    num1_list = clean_numbers(num1).split(",")
    num2_list = clean_numbers(num2).split(",")
    for i in range(0,len(num1_list)):
        num1_list[i] = num1_list[i][-10:]
    for i in range(0,len(num2_list)):
        num2_list[i] = num2_list[i][-10:]
    combined = num1_list + num2_list
    counter = collections.Counter(combined)
    for i in counter.values():
        if i>1:
            a =a +1
    percentage_list1=(a/len(num1_list))*100
    percentage_list2=(a/len(num2_list))*100
    return percentage_list1,percentage_list2
import xlwt
z_name = []
j_name = []
j_address = []
z_address=[]
z_phone = []
j_phone = []
import openpyxl
book1 = Workbook()
sheet = book1.active
jd = pd.read_csv("restaurants_jd.csv")
zom = pd.read_csv("restaurant_zomato.csv")
def calculate_percentage(jd,zom):
    for i in range(0,310):
        for j in range(0,310):
            np1,np2 = match_sentences(jd.Name[i],zom.Name[j])
            ap1,ap2 = match_sentences(jd.Address[i],zom.Address[j])
            pp1,pp2 = match_numbers(jd.Phone[i],zom.Phone[j])
            j_name.append(np1)
            z_name.append(np2)
            j_address.append(ap1)
            z_address.append(ap2)
            j_phone.append(pp1)
            z_phone.append(pp2)
calculate_percentage(jd,zom)
for i in range(len(j_name)):
    sheet.cell(row=i+1, column=1).value = j_name[i]
    sheet.cell(row=i+1, column=2).value = z_name[i]
    sheet.cell(row=i+1, column=3).value = j_address[i]
    sheet.cell(row=i+1, column=4).value = z_address[i]
    sheet.cell(row=i+1, column=5).value = j_phone[i]
    sheet.cell(row=i+1, column=6).value = z_phone[i]
book1.save("All_Training_Data.xlsx")

